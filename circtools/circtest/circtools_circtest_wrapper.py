#!/usr/bin/env python3

import pandas as pd
import os
from scripts.circtest_functions import circ_filter, circ_test, circ_test_pairwise
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
import matplotlib.pyplot as plt
import matplotlib.backends.backend_pdf as pdf_backend
import numpy as np
import warnings
import matplotlib.patches as mpatches


logger = logging.getLogger(__name__)
MAX_LINES = 50000




# Colour palette scales to any number of groups.
# Colour mode: evenly-spaced hues around the HSL wheel.
# BW mode: evenly-spaced greys from dark to light.

def _make_palette(n_groups, colour_mode):
    """Return a list of n_groups hex colour strings."""
    if colour_mode != "colour":
        # dark → light greyscale
        greys = [int(40 + 180 * i / max(n_groups - 1, 1)) for i in range(n_groups)]
        return [f"#{v:02X}{v:02X}{v:02X}" for v in greys]
    # Qualitative colour wheel via matplotlib tab10 / tab20
    import matplotlib.cm as cm
    cmap = cm.get_cmap("tab10" if n_groups <= 10 else "tab20")
    return [
        "#{:02X}{:02X}{:02X}".format(
            int(r * 255), int(g * 255), int(b * 255)
        )
        for r, g, b, _ in [cmap(i / max(n_groups - 1, 1)) for i in range(n_groups)]
    ]


def circ_ratio_plot(circ_row, linear_row, coord_row, group_indicators,
                    lab_legend, y_axis_range, colour_mode, ax, sig_p=None,
                    pairwise=None):
    """
    Bar plot of circRNA ratio per group with error bars.

    sig_p    : overall FDR-adjusted p-value from circ_test (shown in subtitle).
    pairwise : dict {(g1, g2): p_adj} for this circRNA — draws significance
               brackets between significant pairs when available.
    """
    circ_vals   = np.array(circ_row.values,   dtype=float)
    linear_vals = np.array(linear_row.values, dtype=float)
    total       = circ_vals + linear_vals
    ratio       = np.where(total > 0, circ_vals / total, np.nan)

    groups   = np.array(group_indicators)
    uniq_grp = list(dict.fromkeys(groups))   # preserve config order
    n_grp    = len(uniq_grp)
    palette  = _make_palette(n_grp, colour_mode)

    bar_width   = 0.55
    x_positions = np.arange(n_grp)

    means, sems = [], []
    for g_idx, grp in enumerate(uniq_grp):
        mask      = groups == grp
        grp_ratio = ratio[mask]
        mean_val  = float(np.nanmean(grp_ratio))
        n_valid   = int(np.sum(~np.isnan(grp_ratio)))
        sem_val   = float(np.nanstd(grp_ratio, ddof=1) / np.sqrt(n_valid)) if n_valid > 1 else 0.0
        means.append(mean_val)
        sems.append(sem_val)

        ax.bar(
            x_positions[g_idx], mean_val,
            width=bar_width,
            color=palette[g_idx],
            edgecolor="black",
            linewidth=0.8,
            zorder=2,
        )
        if sem_val > 0:
            ax.errorbar(
                x_positions[g_idx], mean_val,
                yerr=sem_val,
                fmt="none",
                ecolor="black",
                elinewidth=1.2,
                capsize=4,
                capthick=1.2,
                zorder=3,
            )

    # ── axis / title ─────────────────────────────────────────────────────────
    gene  = coord_row.get("Gene", coord_row.get("gene", "unknown"))
    chrom = coord_row.iloc[0]
    start = coord_row.iloc[1]
    end   = coord_row.iloc[2]

    p_str = ""
    if sig_p is not None:
        try:
            sp = float(sig_p)
            p_str = f"  (FDR p = {sp:.2e})" if sp < 0.001 else f"  (FDR p = {sp:.4f})"
        except Exception:
            pass

    ax.set_title(
        f"{gene}  –  Chr{chrom}:{start}-{end}{p_str}",
        fontsize=10, loc="left", pad=6,
    )

    y_max = y_axis_range if y_axis_range > 0 else max(
        (m + s for m, s in zip(means, sems)), default=1.0
    ) * 1.35
    ax.set_ylim(0, y_max)
    ax.set_xlim(-0.6, n_grp - 0.4)

    # Group name labels on x-axis (no need for a separate legend)
    ax.set_xticks(x_positions)
    ax.set_xticklabels(uniq_grp, fontsize=9, rotation=30 if n_grp > 4 else 0,
                       ha="right" if n_grp > 4 else "center")
    ax.set_ylabel("circRNA / (circRNA + linear)", fontsize=9)
    ax.set_xlabel("")

    y_ticks = np.linspace(0, y_max, 5)
    ax.set_yticks(y_ticks)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v:.2f}"))

    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)

    # ── pairwise significance brackets ────────────────────────────────────────
    def _sig_label(p):
        try:
            p = float(p)
        except Exception:
            return "ns"
        if p < 1e-3: return "***"
        if p < 1e-2: return "**"
        if p < 5e-2: return "*"
        return "ns"

    if pairwise:
        grp_idx = {g: i for i, g in enumerate(uniq_grp)}
        sig_pairs = sorted(
            [(pair, p) for pair, p in pairwise.items()
             if _sig_label(p) != "ns"
             and pair[0] in grp_idx and pair[1] in grp_idx],
            key=lambda kv: abs(grp_idx[kv[0][1]] - grp_idx[kv[0][0]])
        )
        if sig_pairs:
            bar_top = max(m + s for m, s in zip(means, sems)) if means else y_max
            step    = (y_max - bar_top) / max(len(sig_pairs) + 1, 2)
            step    = max(step, y_max * 0.05)
            tick_h  = step * 0.3

            for level, ((g1, g2), p_adj) in enumerate(sig_pairs):
                x1, x2 = grp_idx[g1], grp_idx[g2]
                y_br    = bar_top + (level + 1) * step

                ax.plot([x1, x1, x2, x2],
                        [y_br - tick_h, y_br, y_br, y_br - tick_h],
                        color="black", linewidth=0.9)
                ax.text((x1 + x2) / 2, y_br + tick_h * 0.3,
                        _sig_label(p_adj),
                        ha="center", va="bottom", fontsize=9)


def run_circ_test_wrapper(args):

    print("Loading CircRNACount")
    CircRNACount = pd.read_csv(
        os.path.join(args.dcc_data, "CircRNACount"), sep="\t", header=0
    )

    print("Loading LinearCount")
    LinearCount = pd.read_csv(
        os.path.join(args.dcc_data, "LinearCount"), sep="\t", header=0
    )

    print("Loading CircCoordinates")
    CircCoordinates = pd.read_csv(
        os.path.join(args.dcc_data, "CircCoordinates"), sep="\t", header=0
    )

    cond_cols_1based = [int(x) for x in args.condition_columns.split(",")]
    cond_cols = [x - 1 for x in cond_cols_1based]        

    circle_description = list(range(3))                   

    cond_names = args.condition_list.split(",")            
    group_indices = [int(x) for x in args.groups.split(",")] 


    final_grouping = [cond_names[g - 1] for g in group_indices]

    print(f"[INFO] Condition columns (0-based): {cond_cols}")
    print(f"[INFO] Group labels (per sample):   {final_grouping}")


    all_cols = circle_description + cond_cols

    circ_slice  = CircRNACount.iloc[:, all_cols]
    linear_slice = LinearCount.iloc[:, all_cols]


    print("Filtering circRNA counts")
    CircRNACount_filtered = circ_filter(
        circ=circ_slice,
        linear=linear_slice,
        Nreplicates=args.replicates,
        filter_sample=args.filter_sample,
        filter_count=args.filter_count,
        percentage=args.percent_filter,
        circle_description=circle_description
    )


    print("Filtering circRNA coordinates")
    CircCoordinates_filtered = CircCoordinates.loc[CircRNACount_filtered.index]

    print("Filtering linear RNA counts")
    LinearCount_filtered = linear_slice.loc[CircRNACount_filtered.index]

    print("Running circTest")
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", category=RuntimeWarning, 
                                message="invalid value encountered in sqrt")
        results = circ_test(
            Circ=CircRNACount_filtered,
            Linear=LinearCount_filtered,
            CircCoordinates=CircCoordinates_filtered,
            group=final_grouping,
            alpha=args.max_fdr,
            circle_description=circle_description
        )

    summary = results["summary_table"]

    if len(summary) == 0:
        print("No candidates to plot, exiting.")
        return

    # ── Pairwise beta-binomial LRT (only for significant circRNAs) ────────────
    sig_indices = list(summary.index)
    n_groups    = len(set(final_grouping))

    if n_groups > 2 and sig_indices:
        print(f"Running pairwise tests for {len(sig_indices)} significant circRNAs across {n_groups} groups...")
        # circle_description was resolved to names inside circ_test; re-resolve here
        resolved_desc = list(circ_slice.columns[:3])
        pairwise_df = circ_test_pairwise(
            Circ               = CircRNACount_filtered,
            Linear             = LinearCount_filtered,
            group              = final_grouping,
            sig_indices        = sig_indices,
            circle_description = resolved_desc,
            CircCoordinates    = CircCoordinates_filtered,
            alpha              = args.max_fdr,
        )
        pairwise_path = args.output_name + "_pairwise.csv"
        pairwise_df.to_csv(pairwise_path, sep="\t", index=False)
        print(f"Pairwise results saved to {pairwise_path}")
    else:
        if n_groups <= 2:
            print("Only 2 groups — pairwise test is identical to omnibus, skipping.")
        pairwise_df = None


    print("Generating plots")
    max_n = min(args.max_plots, len(summary))

    if args.only_negative:
        plot_rows = [
            i for i in summary.index
            if "direction" in summary.columns and summary.loc[i, "direction"] < 0
        ]
    else:
        plot_rows = list(summary.index[:max_n])

    if plot_rows:
        pdf_path = args.output_name + ".pdf"

        # Build pairwise lookup keyed by row-index integer for fast access
        pairwise_by_idx = {}
        if pairwise_df is not None and not pairwise_df.empty:
            for _, pw_row in pairwise_df.iterrows():
                cid  = pw_row["CircID"]
                pair = (str(pw_row["group1"]), str(pw_row["group2"]))
                p    = float(pw_row["p_adj"]) if not pd.isna(pw_row["p_adj"]) else float("nan")
                pairwise_by_idx.setdefault(cid, {})[pair] = p

        # Build CircID string lookup for the filtered count table
        circ_id_str = {}
        for idx in CircRNACount_filtered.index:
            coord = CircCoordinates_filtered.loc[idx]
            chrom = str(coord.iloc[0])
            start = str(coord.iloc[1])
            end   = str(coord.iloc[2])
            circ_id_str[idx] = f"{chrom}:{start}-{end}"

        with pdf_backend.PdfPages(pdf_path) as pdf:
            for circ_id in plot_rows:
                fig, ax = plt.subplots(1, 1, figsize=(7, 5))

                n_desc     = len(circle_description)
                circ_row   = CircRNACount_filtered.loc[circ_id].iloc[n_desc:]
                linear_row = LinearCount_filtered.loc[circ_id].iloc[n_desc:]
                coord_row  = CircCoordinates_filtered.loc[circ_id]

                # sig_p from summary table
                sig_p = summary.loc[circ_id, "sig_p"] if circ_id in summary.index else None

                # pairwise brackets keyed by CircID string
                cid_str  = circ_id_str.get(circ_id, str(circ_id))
                pairwise = pairwise_by_idx.get(cid_str)

                circ_ratio_plot(
                    circ_row, linear_row, coord_row,
                    group_indicators=final_grouping,
                    lab_legend=args.output_label,
                    y_axis_range=args.y_range,
                    colour_mode=args.colour_mode,
                    ax=ax,
                    sig_p=sig_p,
                    pairwise=pairwise,
                )

                fig.tight_layout()
                pdf.savefig(fig, bbox_inches="tight")
                plt.close(fig)

        print(f"Plots saved to {pdf_path}")

    print("Saving CSV")
    csv_data = summary.head(MAX_LINES).dropna()     
    csv_data.to_csv(
        args.output_name + ".csv",
        sep="\t",
        index=False,
        header=args.add_header
    )

    print("Saving Excel")
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Significant circles"
    sig_df = summary.head(MAX_LINES)
    ws1.append(list(sig_df.columns))
    for row in sig_df.itertuples(index=False):
        ws1.append(list(row))

    sig_idx = sig_df.index
    ws2 = wb.create_sheet("Circle Counts")
    for row in dataframe_to_rows(
        circ_slice.loc[sig_idx], index=False, header=True
    ):
        ws2.append(row)

    ws3 = wb.create_sheet("Linear Counts")
    for row in dataframe_to_rows(
        linear_slice.loc[sig_idx], index=False, header=True
    ):
        ws3.append(row)

    wb.save(args.output_name + ".xlsx")
    print("Done.")