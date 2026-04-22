#!/usr/bin/env python3

import pandas as pd
import argparse
import os
from scripts.circtest_functions import circ_filter, circ_test, summary_se
from openpyxl import Workbook
import matplotlib.pyplot as plt
from openpyxl.utils.dataframe import dataframe_to_rows


import logging
import sys

logger = logging.getLogger(__name__)
MAX_LINES = 50000

def parse_args():
    parser = argparse.ArgumentParser(description="Run CircTest wrapper in Python.")
    parser.add_argument("dcc_data", help="Path to CircRNACount, LinearCount, and CircCoordinates files")
    parser.add_argument("replicates", type=int)
    parser.add_argument("condition_list", help="Comma-separated condition names")
    parser.add_argument("condition_columns", help="Comma-separated column indices (1-based)")
    parser.add_argument("output_name")
    parser.add_argument("max_fdr", type=float)
    parser.add_argument("max_plots", type=int)
    parser.add_argument("filter_sample", type=int)
    parser.add_argument("filter_count", type=int)
    parser.add_argument("groups", help="Comma-separated list of group numbers")
    parser.add_argument("output_label")
    parser.add_argument("percent_filter", type=float)
    parser.add_argument("only_negative", type=bool)
    parser.add_argument("add_header", type=bool)
    parser.add_argument("y_range", type=float)
    parser.add_argument("colour_mode", choices=["colour", "bw"])
    return parser.parse_args()

def run_circ_test_wrapper(args):

    print("Loading CircRNACount")
    circ_file = os.path.join(args.dcc_data, "CircRNACount")
    CircRNACount = pd.read_csv(circ_file, sep="\t", header=0)

    print("Loading LinearCount")
    linear_file = os.path.join(args.dcc_data, "LinearCount")
    LinearCount = pd.read_csv(linear_file, sep="\t", header=0)

    print("Loading CircCoordinates")
    coord_file = os.path.join(args.dcc_data, "CircCoordinates")
    CircCoordinates = pd.read_csv(coord_file, sep="\t", header=0)

    # Parse condition columns and group info
        # Parse condition columns and group info
    cond_cols = [int(x) - 1 for x in args.condition_columns.split(",")]
    cond_names = args.condition_list.split(",")

    # Map group indices (1-based) to condition names
    group_indices = [int(x) for x in args.groups.split(",")]
    if any((g < 1 or g > len(cond_names)) for g in group_indices):
        raise IndexError(f"Invalid group index in -g: {group_indices} (valid 1–{len(cond_names)})")

    # Final mapped group labels (e.g., ['treatment', 'control', 'control'])
    group_vals = [cond_names[i - 1] for i in group_indices]

    print(f"[INFO] Group labels (mapped): {group_vals}")

    # Determine metadata (non-count) columns dynamically
    circle_description = list(range(min(cond_cols)))


    print(f"Using circle description columns: {circle_description}")
    print("Filtering circRNA counts")
    CircRNACount_filtered = circ_filter(
        circ=CircRNACount.iloc[:, circle_description + cond_cols],
        linear=LinearCount.iloc[:, circle_description + cond_cols],
        Nreplicates=args.replicates,
        filter_sample=args.filter_sample,
        filter_count=args.filter_count,
        percentage=args.percent_filter,
        circle_description=circle_description
    )

    print("Filtering circRNA coordinates")
    CircCoordinates_filtered = CircCoordinates.loc[CircRNACount_filtered.index]
    print(f"[DEBUG] circ_filter returned {CircRNACount_filtered.shape[0]} rows")
    print(f"[DEBUG] CircRNACount_filtered shape before circ_test: {CircRNACount_filtered.shape}")



    print("Filtering linear RNA counts")
    LinearCount_filtered = LinearCount.loc[CircRNACount_filtered.index, CircRNACount_filtered.columns]

    print("Running circTest")
    results = circ_test(
        Circ=CircRNACount_filtered,
        Linear=LinearCount_filtered,
        CircCoordinates=CircCoordinates_filtered,
        group=group_vals,
        alpha=args.max_fdr,
        circle_description=circle_description
    )

    if len(results["summary_table"]) == 0:
        print("No candidates to plot, exiting.")
        return

    max_n = min(args.max_plots, len(results["summary_table"]))

    print("Generating plots")
    for i, circ_id in enumerate(results["summary_table"].index[:max_n]):
        # Placeholder for future plot function
        pass

    print("Saving CSV")
    results["summary_table"].head(MAX_LINES).to_csv(
        args.output_name + ".csv", sep="\t", index=False, header=args.add_header
    )
    
    print("Saving Excel")
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Significant circles"

    # Add custom header
    header = [
        "Chr", "Start", "End", "Gene", "JunctionType",
        "Strand", "Start.End.Region", "OverallRegion",
        "sig_p", "raw_p", "group_1_ratio_mean", "group_2_ratio_mean"
    ]

    ws1.append(header)

    # Add the data rows
    for row in pd.DataFrame(results["summary_table"].head(MAX_LINES)).itertuples(index=False):
        ws1.append(list(row))

    # Add Circle Counts sheet
    wb.create_sheet("Circle Counts")
    for row in dataframe_to_rows(CircRNACount_filtered.loc[results["summary_table"].index], index=False, header=True):
        wb["Circle Counts"].append(row)

    # Add Linear Counts sheet
    wb.create_sheet("Linear Counts")
    for row in dataframe_to_rows(LinearCount_filtered.loc[results["summary_table"].index], index=False, header=True):
        wb["Linear Counts"].append(row)

    wb.save(args.output_name + ".xlsx")

if __name__ == "__main__":
    args = parse_args()
    run_circ_test_wrapper(args)
